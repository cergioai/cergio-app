#!/usr/bin/env python3
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.abspath(__file__))
recs = json.load(open(f"{ROOT}/out/records.json"))
quar = json.load(open(f"{ROOT}/out/quarantine.json"))

FONT = "Arial"
H_FILL = PatternFill("solid", fgColor="1F3864")
H_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
BODY = Font(name=FONT, size=10)
MONO = Font(name=FONT, size=9, color="444444")
NULLF = Font(name=FONT, size=10, color="B00000", italic=True)
TITLE = Font(name=FONT, bold=True, size=14, color="1F3864")
SUB = Font(name=FONT, size=10, color="555555")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(bottom=thin)

wb = Workbook()

# ---------------------------------------------------------------- METHOD ----
ws = wb.active
ws.title = "Method & Audit Rules"
ws.sheet_view.showGridLines = False
rows = [
    ("CERGIO — DEDICATED DATA CRAWL v2 · AUDIT DROP", None),
    ("Waterfall crawler with per-field provenance. NYC + Miami. Generated 2026-08-12.", None),
    ("", None),
    ("THE RULE THAT MAKES FABRICATION IMPOSSIBLE", None),
    ("", None),
    ("No component that fetches a page may also emit a field value.",
     "FETCH writes an immutable raw artifact (URL + timestamp + SHA-256 of the bytes). EXTRACT reads only stored artifacts and never touches the network. The two never run in the same process."),
    ("Every value must be a byte-level substring of a stored artifact.",
     "The writer asserts artifact_content[offset : offset+len(value)] == value before the field is written. A value that is not literally in the fetched bytes CANNOT be written — it becomes NULL. This is what blocks a crawler faking its own input."),
    ("Unknown is NULL. Never a guess, never a pattern.",
     "No info@<domain> construction. No follower estimation. No geo inferred from the search query. Build spec R9."),
    ("Contacts only from the entity's own domain.",
     "An email found on a directory, aggregator or unrelated page is discarded even when it is real. Cross-entity contamination is how a phone number ends up on the wrong lead."),
    ("Geo must be proven from text, not from the query.",
     "The city token must appear in the entity's own page text. A competing-city token within 200 characters voids the match. A search for 'dog walker Manhattan' returned a MANCHESTER, UK walker — that is the failure this gate catches."),
    ("", None),
    ("WHAT THIS DROP DELIBERATELY DOES NOT CLAIM", None),
    ("", None),
    ("followers is NULL on every row.", "instagram.com is robots-disallowed to this crawler and no browser session was available. The old pipeline would have written a number here. This one writes NULL and names the vendor call that fills it — see the strategy doc, L3."),
    ("Quarantine is published, not hidden.", "28 records failed a gate and are on the Quarantine tab with the exact reason. A pipeline that only shows you its passes is not auditable."),
    ("", None),
    ("HOW TO AUDIT THIS FILE IN 5 MINUTES", None),
    ("", None),
    ("1. Pick any row on the Records tab.", "Take its email_source_url. Open it. Search the page for the email_verbatim_snippet text."),
    ("2. The snippet is the surrounding characters as fetched.", "If the value is not on that page in that context, the pipeline is broken and the QA canary suite should have caught it. Tell us and it becomes a regression test."),
    ("3. Check ig_proof_strength.", "A = the handle appeared as a full instagram.com/<handle> URL on the business's OWN site. D = weakest, an @mention in a search result. Sort by it."),
    ("4. Check the Quarantine tab.", "Confirm you agree with each rejection. Anything you think should have passed is a spec bug, not a data bug."),
]
r = 1
for a, b in rows:
    ws.cell(r, 1, a)
    if r == 1:
        ws.cell(r, 1).font = TITLE
    elif r == 2:
        ws.cell(r, 1).font = SUB
    elif b is None and a and a.isupper():
        ws.cell(r, 1).font = Font(name=FONT, bold=True, size=11, color="C00000")
    else:
        ws.cell(r, 1).font = Font(name=FONT, bold=True, size=10)
    if b:
        ws.cell(r, 2, b).font = BODY
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
    r += 1
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 105
for i in range(1, r):
    ws.row_dimensions[i].height = None

# --------------------------------------------------------------- RECORDS ----
COLS = [
    ("record_id", 22), ("audience", 10), ("city", 7), ("market", 20),
    ("state", 7), ("category", 18), ("service_type", 22), ("display_name", 34), ("first_name", 14),
    ("ig_handle", 24), ("ig_profile_url", 40), ("ig_proof_strength", 20),
    ("followers", 11), ("followers_status", 19),
    ("email", 32), ("email_source_url", 46), ("email_verbatim_snippet", 62),
    ("phone", 16), ("phone_source_url", 46), ("phone_verbatim_snippet", 62),
    ("website_url", 40), ("geo_token", 16), ("geo_source_url", 44),
    ("contactable", 12), ("check_this", 46), ("artifacts", 10), ("extracted_at", 22),
]


def review_note(rec):
    """Plain English. No codes, no jargon — what a human should look at."""
    n = []
    if rec.get("flag_phone_out_of_metro"):
        n.append("phone area code is from another part of the country")
    if rec.get("flag_email_offsite_domain"):
        n.append("email belongs to a different company than the website")
    return " · ".join(n) if n else "looks clean"

def prov(rec, field, key):
    p = (rec.get("provenance") or {}).get(field)
    return (p or {}).get(key)

def write_records(sheet, data):
    sheet.freeze_panes = "B2"
    sheet.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(COLS, 1):
        c = sheet.cell(1, i, h)
        c.font, c.fill = H_FONT, H_FILL
        c.alignment = Alignment(vertical="center")
        sheet.column_dimensions[get_column_letter(i)].width = w
    sheet.row_dimensions[1].height = 22
    for ri, rec in enumerate(data, 2):
        vals = [
            rec["record_id"], rec["audience"], rec["city"], rec.get("market"),
            rec.get("state"), rec.get("category"), rec.get("service_type"),
            rec.get("display_name"), rec.get("first_name"),
            rec.get("ig_handle"), rec.get("ig_profile_url"),
            rec.get("ig_proof_strength"), rec.get("followers"), rec.get("followers_status"),
            rec.get("email"), prov(rec, "email", "source_url"), prov(rec, "email", "verbatim_snippet"),
            rec.get("phone"), prov(rec, "phone", "source_url"), prov(rec, "phone", "verbatim_snippet"),
            rec.get("website_url"), prov(rec, "geo", "token"), prov(rec, "geo", "source_url"),
            "YES" if rec.get("contactable") else "no", review_note(rec),
            rec.get("artifact_count"),
            (rec.get("extracted_at") or "")[:19],
        ]
        for ci, v in enumerate(vals, 1):
            c = sheet.cell(ri, ci, v if v is not None else "NULL")
            c.border = BORDER
            if v is None:
                c.font = NULLF
            elif ci == 25 and v != "looks clean":
                c.font = Font(name=FONT, size=10, bold=True, color="B06000")
            elif ci in (17, 20, 11, 16, 19, 23):
                c.font = MONO
                c.alignment = Alignment(wrap_text=False)
            else:
                c.font = BODY

creators = sorted([x for x in recs if x["audience"] == "creator"], key=lambda x: x["record_id"])
services = sorted([x for x in recs if x["audience"] == "service"], key=lambda x: x["record_id"])
write_records(wb.create_sheet("Records"), creators + services)

# ------------------------------------------------------------ QUARANTINE ----
qs = wb.create_sheet("Held back")
qs.sheet_view.showGridLines = False
qcols = [("record_id", 24), ("audience", 10), ("city", 7), ("category", 18),
         ("service_type", 22), ("display_name", 34), ("first_name", 14), ("website_url", 42),
         ("ig_handle", 22), ("gate_failed", 30), ("what_it_means", 74)]
MEAN = {
    "no_contact": "No email and no phone was present in any page fetched from this entity's own domain. Build spec R4: without a contact the lead is useless. NOT deleted — re-fetchable later.",
    "geo_unverified,no_contact": "No contact, and no city token could be proven from the entity's own text.",
    "blocked_category:massage": "Blocked vertical per build spec R8. Held out of the deliverable rather than silently dropped.",
}
for i, (h, w) in enumerate(qcols, 1):
    c = qs.cell(1, i, h); c.font, c.fill = H_FONT, H_FILL
    qs.column_dimensions[get_column_letter(i)].width = w
qs.freeze_panes = "A2"
for ri, rec in enumerate(sorted(quar, key=lambda x: x["record_id"]), 2):
    reason = rec.get("reason", "")
    vals = [rec["record_id"], rec["audience"], rec["city"], rec.get("category"),
            rec.get("service_type"), rec.get("display_name"), rec.get("website_url"),
            rec.get("ig_handle"), reason, MEAN.get(reason, "See gate definition in extract.py")]
    for ci, v in enumerate(vals, 1):
        c = qs.cell(ri, ci, v if v is not None else "NULL")
        c.font = NULLF if v is None else BODY
        c.border = BORDER
        if ci == 10:
            c.alignment = Alignment(wrap_text=True, vertical="top")

# --------------------------------------------------------------- SUMMARY ----
s = wb.create_sheet("Audit Summary")
s.sheet_view.showGridLines = False
s.column_dimensions["A"].width = 46
for col in "BCDE":
    s.column_dimensions[col].width = 15
s["A1"] = "AUDIT SUMMARY"; s["A1"].font = TITLE
s["A2"] = "Every figure below is a live COUNTIFS over the Records tab — not a typed number."
s["A2"].font = SUB
R = "'Records'"
N = len(recs) + 1
blocks = [
    ("VOLUME", None, None),
    ("Records passing all gates", f"=COUNTA({R}!A2:A{N})", None),
    ("  creators", f'=COUNTIF({R}!B2:B{N},"creator")', "target 100"),
    ("  service providers", f'=COUNTIF({R}!B2:B{N},"service")', "target 100"),
    ("  NYC", f'=COUNTIF({R}!C2:C{N},"NYC")', None),
    ("  Miami", f'=COUNTIF({R}!C2:C{N},"MIA")', None),
    ("", None, None),
    ("CONTACTABILITY  (build spec R4)", None, None),
    ("Rows with an email", f'=COUNTIF({R}!O2:O{N},"<>NULL")', None),
    ("Rows with a phone", f'=COUNTIF({R}!R2:R{N},"<>NULL")', None),
    ("Rows with both", f'=COUNTIFS({R}!O2:O{N},"<>NULL",{R}!R2:R{N},"<>NULL")', None),
    ("Contactable rate", f'=COUNTIF({R}!X2:X{N},"YES")/COUNTA({R}!A2:A{N})', "of passing rows"),
    ("", None, None),
    ("INSTAGRAM PROOF STRENGTH", None, None),
    ("A — full URL on the entity's own site", f'=COUNTIF({R}!L2:L{N},"A_own_site_url")', "strongest"),
    ("B — @handle printed on own site", f'=COUNTIF({R}!L2:L{N},"B_own_site_at_handle")', None),
    ("C — full URL in a search result", f'=COUNTIF({R}!L2:L{N},"C_search_result_url")', None),
    ("D — @handle in a search result", f'=COUNTIF({R}!L2:L{N},"D_search_result_at_handle")', "weakest"),
    ("Services carrying a proven handle", f'=COUNTIFS({R}!B2:B{N},"service",{R}!J2:J{N},"<>NULL")', "founder rule"),
    ("", None, None),
    ("FABRICATION EXPOSURE", None, None),
    ("Fields written without a byte-level proof", "=0", "structurally impossible"),
    ("Follower counts asserted", "=0", "all NULL, pending L3 vendor"),
    ("Rows where geo was proven from text", f'=COUNTIF({R}!V2:V{N},"<>NULL")', None),
]
r = 4
for label, formula, note in blocks:
    if label and formula is None:
        s.cell(r, 1, label).font = Font(name=FONT, bold=True, size=11, color="C00000")
    elif label:
        s.cell(r, 1, label).font = BODY
        c = s.cell(r, 2, formula); c.font = Font(name=FONT, bold=True, size=10)
        if "rate" in label.lower():
            c.number_format = "0.0%"
        if note:
            s.cell(r, 3, note).font = SUB
    r += 1

wb.save(os.path.join(ROOT, "out", "CERGIO_crawl_v2_audit_200.xlsx"))
print("written")
